from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook
from PIL import Image

from jpg2xlsx import convert


def make_image(path: Path) -> None:
    image = Image.new("RGB", (2, 2))
    image.putdata(
        [
            (255, 0, 0),
            (0, 255, 0),
            (0, 0, 255),
            (255, 255, 0),
        ]
    )
    image.save(path)


def test_convert_appends_suffix_and_writes_workbook(tmp_path: Path) -> None:
    input_path = tmp_path / "source.png"
    output_path = tmp_path / "result"
    make_image(input_path)

    convert(input_path, output_path)

    workbook_path = tmp_path / "result.xlsx"
    assert workbook_path.exists()

    workbook = load_workbook(workbook_path)
    worksheet = workbook.active
    assert worksheet.max_row == 2
    assert worksheet.max_column == 2
    assert worksheet["A1"].fill.fgColor.rgb == "FFFF0000"
    assert worksheet["B1"].fill.fgColor.rgb == "FF00FF00"
    assert worksheet["A2"].fill.fgColor.rgb == "FF0000FF"
    assert worksheet["B2"].fill.fgColor.rgb == "FFFFFF00"


def test_convert_creates_parent_directories(tmp_path: Path) -> None:
    input_path = tmp_path / "source.png"
    output_path = tmp_path / "nested" / "output.xlsx"
    make_image(input_path)

    convert(input_path, output_path)

    assert output_path.exists()


def test_convert_can_downscale_large_images(tmp_path: Path) -> None:
    input_path = tmp_path / "source.png"
    output_path = tmp_path / "scaled.xlsx"
    image = Image.new("RGB", (8, 4), (12, 34, 56))
    image.save(input_path)

    convert(input_path, output_path, max_dimension=4)

    workbook = load_workbook(output_path)
    worksheet = workbook.active
    assert worksheet.max_column == 4
    assert worksheet.max_row == 2


def test_convert_can_quantize_colors(tmp_path: Path) -> None:
    input_path = tmp_path / "source.png"
    output_path = tmp_path / "quantized.xlsx"
    image = Image.new("RGB", (4, 1))
    image.putdata(
        [
            (255, 0, 0),
            (250, 10, 10),
            (0, 255, 0),
            (10, 250, 10),
        ]
    )
    image.save(input_path)

    convert(input_path, output_path, colors=2)

    with ZipFile(output_path) as workbook_archive:
        styles_xml = workbook_archive.read("xl/styles.xml").decode()

    assert styles_xml.count("patternFill") < 8
